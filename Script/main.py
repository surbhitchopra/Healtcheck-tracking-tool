import sys
import re
import csv
from pathlib import Path
from datetime import datetime
import openpyxl as opxl
import hcfuncs as funcs
import time
import socket
import gc  # Garbage collection for memory management

# COMPREHENSIVE PERFORMANCE TRACKING
script_start_time = time.time()
step_times = {}

def log_step_time(step_name, start_time):
    """Log the time taken for each major step"""
    duration = time.time() - start_time
    step_times[step_name] = duration
    if duration > 30:  # Log steps taking more than 30 seconds
        print(f"\n!!! SLOW STEP: {step_name} took {duration:.2f}s")
    return time.time()  # Return new start time

print(f">> OPTIMIZED 1830PSS Health Check Processing Started for ALL NETWORKS...")
print(f">> Performance monitoring enabled - will flag steps >30s")

# Set paths for directories and paths

current_dir = Path.cwd()
input_hc_dir = current_dir / Path(r"input-hc-report")
output_dir = current_dir / Path(r"output")

# Delete all files in the output_dir (with Windows file locking protection)
import time
for item in output_dir.iterdir():
    if item.is_file():
        for attempt in range(3):  # Try up to 3 times
            try:
                item.unlink()
                break  # Success, exit retry loop
            except PermissionError as e:
                if attempt == 2:  # Last attempt
                    print(f"Warning: Could not delete {item.name}: {e}")
                    print(f"File may be open in Excel or another application")
                    # Continue processing instead of failing
                else:
                    time.sleep(1)  # Wait 1 second before retry
            except Exception as e:
                print(f"Warning: Could not delete {item.name}: {e}")
                break  # Don't retry for other exceptions

# List all files in input_hc_dir
dirList = [file for file in input_hc_dir.iterdir() if file.is_file()]

# Check if the HC report and the remote inventory CSV are loaded.
input_files_message = '\nPlease ensure that the HC report workbook and the remote inventory CSV are loaded properly.\n'

if(len(dirList)) != 2:
    print(input_files_message)
    sys.exit()

# Assign filenames to hc report and remote inventory
for file in dirList:
    if file.suffix == '.xlsx':
        hc_filename = file.name
    elif file.suffix == '.csv':
        rem_inv_filename = file.name

# Print the filenames of the hc report and remote inventory
print(f"\nThe filename of the HC report being analyzed is {hc_filename}")
print(f"The filename of the Remote Inventory being analyzed is {rem_inv_filename}\n")

# Get the network_name, hc_report_year, hc_report_month and hc_report_date from the hc_filename

# Extract network name to match your exact file naming
# Your file: BSNL_West_Zone_DWDM_Reports_20250809.xlsx
# Should create: BSNL_West_Zone_DWDM_HC_Issues_Tracker.xlsx
filename_parts = hc_filename.replace('.xlsx', '').split('_')
print(f"Filename parts: {filename_parts}")

# Find where "Reports" appears and take everything before it
if 'Reports' in filename_parts:
    reports_index = filename_parts.index('Reports')
    network_name = '_'.join(filename_parts[:reports_index])
else:
    # Fallback: assume last part is date, second last might be "Reports"
    # Take everything except last 2 parts as network name
    if len(filename_parts) >= 3:
        network_name = '_'.join(filename_parts[:-2])
    else:
        # If filename format is unexpected, use first part
        network_name = filename_parts[0]

# Keep the original case for network name (don't convert to lowercase)
network_name = network_name.strip()
print(f"Network name extracted: {network_name}")

hc_report_year_position = hc_filename.rfind("_") + 1
hc_report_year = hc_filename[hc_report_year_position : hc_report_year_position + 4]
hc_report_month_position = hc_filename.rfind("_") + 5
hc_report_month = hc_filename[hc_report_month_position : (hc_report_month_position + 2)]
hc_report_date_position = hc_filename.rfind("_") + 7
hc_report_date = hc_filename[hc_report_date_position : (hc_report_date_position + 2)]
year_month = hc_report_year + hc_report_month
year_month_date = hc_report_year + hc_report_month + hc_report_date


print(f"Network name is {network_name}\n")
print(f"HC report year is {hc_report_year}")
print(f"HC report month is {hc_report_month}")
print(f"HC report date is {hc_report_date}\n")

# Load the HC report workbook in read mode and activate the CWBP worksheet and Network summary sheet
hagen_report = opxl.load_workbook(input_hc_dir / Path(hc_filename), read_only=False)
cwbp_sheet = hagen_report["CWBP"]
network_summary_sheet = hagen_report["Network Report Summary"]

# Get the maximum number of rows in hc workbook CWBP sheet
cwbp_max_row = cwbp_sheet.max_row

# Get the maximum number of columns in hc workbook CWBP sheet
cwbp_max_col = cwbp_sheet.max_column

# Create the report workbook to store our test cases
hc_test_cases_report = opxl.Workbook()
report_sheet = hc_test_cases_report.active
report_sheet.title = "W & F"

# Print the column titles in the report
for i in range(1, cwbp_max_col + 1):
    report_sheet.cell(row=1, column=i).value = cwbp_sheet.cell(row=1, column=i).value

# Save the workbook as 'extracted_hc_test_cases.xlsx' in output_dir
hc_test_cases_report.save(output_dir / Path(r"extracted_hc_test_cases.xlsx"))

print("Extracting all FAILURES and WARNINGS from Hagen's HC report...", end="")

# Extract all warnings and failures of interest and copy to report sheet
for row in cwbp_sheet.iter_rows(min_row=2, values_only=True):
    # hc_date = row[5][:6]      changed from main-008.py on 2nd Aug 2024
    # if hc_date == year_month: changed from main-008.py on 2nd Aug 2024
    report_sheet.append(row)

# Added on 24th April 2024
# Format the W & F sheet of the extracted hc test cases W & F sheet's HC Id column only
funcs.extracted_sheet_format_hc_id_column(report_sheet, 1)

# Save the workbook as 'extracted_hc_test_cases.xlsx' in output_dir
hc_test_cases_report.save(output_dir / Path(r"extracted_hc_test_cases.xlsx"))

print("Done")

print("Removing the test cases to be ignored from Hagen's HC report...", end="")

#######################################################################################################
# ADDED ON 25TH APRIL 2024 - To filter out the ignored cases from the extracted report
# Started modifying on 29th April 2024 as a three step process

# Open the extracted HC issues for this network and activate the W&F sheet
extracted = output_dir / Path(r"extracted_hc_test_cases.xlsx")
extracted_tracker = opxl.load_workbook(extracted)
extracted_sheet_wf = extracted_tracker["W & F"]
# extracted_sheet_info = extracted_tracker["INFO"]

# Step 1 Remove all 'Info' cases and copy to list
remove_info_from_extracted = [
    row
    for row in extracted_sheet_wf.iter_rows(min_row=2, values_only=True)
    if row[9] != "Info"
]

# Step 2 Remove all the ignored test cases for the network as given that network's ignore text file
# Read-In the test cases to be ignored - with case-insensitive search
ignored_text_file = current_dir / Path(network_name + "_ignored_test_cases.txt")

# Smart file search: Try exact name first, then case-insensitive search
found_text_file = None
if ignored_text_file.exists():
    found_text_file = ignored_text_file
else:
    # Case-insensitive search for existing files
    expected_name_lower = ignored_text_file.name.lower()
    for file in current_dir.glob("*ignored_test_cases.txt"):
        if file.name.lower() == expected_name_lower:
            found_text_file = file
            print(f"Found case-variant file: {file.name} (expected: {ignored_text_file.name})")
            break

# Check if ignored test case file exists
if found_text_file:
    with open(found_text_file) as f:
        ignored_text_file_cases = f.readlines()
else:
    print(f"Warning: Ignored test case file not found: {ignored_text_file.name}")
    print(f"Auto-creating empty ignored test cases file...")
    # Create empty ignored text file with exact name
    with open(ignored_text_file, 'w') as f:
        f.write("# Add test cases to ignore, one per line\n")
    print(f"Created: {ignored_text_file.name}")
    ignored_text_file_cases = []

# Strip all combinations of line endings - all combinations of \r and \n
ignored_text_file_cases = [line.rstrip("\r\n") for line in ignored_text_file_cases]

# Trim leading and trailing spaces in each test case
ignored_text_file_cases = [line.strip() for line in ignored_text_file_cases]

# Remove blank lines in ignored cases
ignored_text_file_cases = [tc for tc in ignored_text_file_cases if tc != ""]

# Now remove all these cases from the extracted list
ignored_cases = [
    tc for tc in remove_info_from_extracted if tc[7] not in ignored_text_file_cases
]

# Create a workbook to save our ignored_cases list to an excel file set_1_2_extracted_test_cases.xlsx
two_step_extracted_report = opxl.Workbook()
two_step_extracted_sheet = two_step_extracted_report.active
two_step_extracted_sheet.title = "2 Step Extracted"

# Print the first row header
funcs.copy_first_row_hc_tracker(two_step_extracted_sheet, extracted_sheet_wf)

# Append all the ignored_cases to this workbook
for line in ignored_cases:
    two_step_extracted_sheet.append(line)

# Save the two_step_extracted_report in the out directory as  two_step_extracted.xlsx
two_step_extracted_report.save(output_dir / Path(r"two_step_extracted.xlsx"))

# Open the ignored HC issues for this network and activate the MAIN sheet
ignored_cases_filename = network_name + "_ignored_test_cases.xlsx"
ignored_cases_file_path = current_dir / Path(ignored_cases_filename)

# Smart file search for ignored Excel tracker - case-insensitive search and auto-creation
found_excel_file = None
if ignored_cases_file_path.exists():
    found_excel_file = ignored_cases_file_path
else:
    # Case-insensitive search for existing files
    expected_excel_name_lower = ignored_cases_file_path.name.lower()
    for file in current_dir.glob("*ignored_test_cases.xlsx"):
        if file.name.lower() == expected_excel_name_lower:
            found_excel_file = file
            print(f"Found case-variant Excel file: {file.name} (expected: {ignored_cases_filename})")
            break

# Load or create ignored Excel tracker
if found_excel_file:
    ignored_cases_workbook = opxl.load_workbook(found_excel_file)
    ignored_sheet = ignored_cases_workbook["MAIN"]
else:
    # Try auto-creation from template
    template_ignored_excel = current_dir / Path("Template_ignored_test_cases .xlsx")
    if template_ignored_excel.exists():
        print(f"\nIgnored test cases Excel tracker not found: {ignored_cases_filename}")
        print(f"Auto-creating from template with EXACT name: {ignored_cases_filename}")
        print(f"Using template: {template_ignored_excel.name}")
        # Copy template to exact required name
        import shutil
        shutil.copy2(template_ignored_excel, ignored_cases_file_path)
        print(f"Created: {ignored_cases_filename}")
        ignored_cases_workbook = opxl.load_workbook(ignored_cases_file_path)
        ignored_sheet = ignored_cases_workbook["MAIN"]
    else:
        print(f"Warning: Ignored test cases Excel tracker not found: {ignored_cases_filename}")
        print(f"Template file also not found: Template_ignored_test_cases .xlsx")
        print(f"Processing will continue with empty ignored cases list...")
        # Create a minimal in-memory ignored cases structure
        class DummyWorkbook:
            def __init__(self):
                self.data = [["HC Date", "HC Id", "Node IP", "Location", "User Label", "NE Type", "Network Name", "Test Case", "Category", "Issue", "Finding", "Task", "Status", "Int/Ext", "Fault Category", "HW Type", "Card Sl No.", "Remarks"]]
            
            def iter_rows(self, min_row=1, values_only=True):
                for i, row in enumerate(self.data):
                    if i + 1 >= min_row:
                        yield row if values_only else [type('Cell', (), {'value': cell}) for cell in row]
        
        class DummySheet:
            def __init__(self):
                self.main = DummyWorkbook()
                
            def __getitem__(self, key):
                return self.main
        
        ignored_cases_workbook = DummySheet()
        ignored_sheet = ignored_cases_workbook["MAIN"]

indices_to_check = (0, 7, 9, 11)  # HC Id, Test Case, Priority and Finding

# Call the function to filter the data
filtered_data, ignored_rows = funcs.remove_ignore_from_extracted(
    two_step_extracted_sheet, ignored_sheet, indices_to_check
)

# Added to main-008.py on 29th July 2024
filtered_data.sort(key=lambda x: x[5])


# Create a new workbook to store our filtered test cases
filtered_test_cases_report = opxl.Workbook()
filtered_sheet = filtered_test_cases_report.active
filtered_sheet.title = "Filtered"

# Print the column titles in the filtered report
for i in range(1, extracted_sheet_wf.max_column + 1):
    filtered_sheet.cell(row=1, column=i).value = extracted_sheet_wf.cell(
        row=1, column=i
    ).value

# Append the filtered_data to this worksheet
for row in filtered_data:
    filtered_sheet.append(row)

# Save the filtered workbook in output_dir
filtered_test_cases_report.save(
    output_dir / Path(r"filtered_from_extracted_hc_test_cases.xlsx")
)

# Create a new workbook to store our ignored test cases
ignored_test_cases_report = opxl.Workbook()
ignored_rows_sheet = ignored_test_cases_report.active
ignored_rows_sheet.title = "Ignored"

# Print the column titles in the ignored test cases report
for i in range(1, extracted_sheet_wf.max_column + 1):
    ignored_rows_sheet.cell(row=1, column=i).value = extracted_sheet_wf.cell(
        row=1, column=i
    ).value

# Append the ignored_rows to this worksheet
for row in ignored_rows:
    ignored_rows_sheet.append(row)

# Save the ignored workbook in output_dir
ignored_test_cases_report.save(
    output_dir / Path(r"ignored_from_extracted_hc_test_cases.xlsx")
)

print("Done")


##########################################################################################################
# Open the master tracker for this network - use exact case from HC filename
hc_issues_tracker_filename = network_name + "_HC_Issues_Tracker.xlsx"
hc_issues_tracker = current_dir / Path(hc_issues_tracker_filename)
print(f"Looking for HC Issues Tracker: {hc_issues_tracker_filename}")

# Use ANY uploaded HC tracker file - same logic for all customers
found_tracker_file = None
if hc_issues_tracker.exists():
    found_tracker_file = hc_issues_tracker
    print(f"Found exact HC tracker: {hc_issues_tracker_filename}")
else:
    # Search for ANY HC tracker file and use the most recent one
    print(f"\nExact tracker not found: {hc_issues_tracker_filename}")
    print(f"Searching for most recent HC tracker file...")
    
    tracker_files = []
    
    # Look for HC tracker patterns
    for pattern in ["*HC*Tracker*.xlsx", "*hc*tracker*.xlsx", "*HC*Issues*.xlsx", "*tracker*.xlsx"]:
        for file in current_dir.glob(pattern):
            if file.name != "Template_HC_Issues_Tracker.xlsx":  # Skip template
                tracker_files.append(file)
    
    # Remove duplicates and sort by modification time (most recent first)
    tracker_files = list(set(tracker_files))
    if tracker_files:
        tracker_files.sort(key=lambda f: f.stat().st_mtime, reverse=True)
        found_tracker_file = tracker_files[0]
        print(f"Found most recent HC tracker: {found_tracker_file.name}")
        print(f"Modified: {datetime.fromtimestamp(found_tracker_file.stat().st_mtime)}")

# Use found file or exit - NO AUTO-CREATION
if not found_tracker_file:
    print(f"\nERROR: No HC tracker file found!")
    print(f"Please upload ANY HC tracker file with a name containing:")
    print(f"   - 'HC' and 'Tracker' (e.g., Any_Network_HC_Issues_Tracker.xlsx)")
    print(f"   - Or 'tracker' (e.g., my_hc_tracker.xlsx)")
    print(f"\nCurrently available xlsx files:")
    for file in current_dir.glob("*.xlsx"):
        print(f"   - {file.name}")
    print(f"\nScript will NOT auto-create files. Please upload your tracker file.")
    sys.exit(1)
else:
    hc_issues_tracker = found_tracker_file
    print(f"Using HC tracker: {found_tracker_file.name}")

master_tracker = opxl.load_workbook(hc_issues_tracker)
master_sheet_main = master_tracker["MAIN"]
master_sheet_open = master_tracker["OPEN"]
master_sheet_closed = master_tracker["CLOSED"]
master_sheet_ignored = master_tracker["IGNORED"]
master_sheet_summary = master_tracker["Summary"]  # Added on 5th Nov 2024
node_coverage_sheet = master_tracker["NODE COVERAGE"]

print("Printing node coverage....")
for row in node_coverage_sheet.iter_rows(min_row=2, values_only=True):
    print(row[0]," ",row[1]," ",row[2]," ",row[3])

funcs.delete_hc_issues_chart(master_sheet_summary)
funcs.delete_node_coverage_chart(master_sheet_summary)

# Copy the first row in HC tracker to the rest three sheets
funcs.copy_first_row_hc_tracker(master_sheet_open, master_sheet_main)
funcs.copy_first_row_hc_tracker(master_sheet_closed, master_sheet_main)
funcs.copy_first_row_hc_tracker(master_sheet_ignored, master_sheet_main)

# Update missing PSS Type in node coverage sheet
nw_summ_sh_data_dict = funcs.update_pss_type(network_summary_sheet, node_coverage_sheet)

# Add new nodes found in the current TEC HC report
funcs.add_new_nodes(nw_summ_sh_data_dict, node_coverage_sheet)

# Added on 10th May 2024 for NE Type
ne_type_dict = {}
for row in node_coverage_sheet.iter_rows(min_row=2, values_only=True):
    ne_type_dict[row[0]] = row[4]


extracted_sheet_wf = extracted_tracker["W & F"]
# extracted_sheet_info = extracted_tracker["INFO"]

# Copy the ignored test cases to the ignored sheet of the master tracker
for row in ignored_rows:
    to_add_in_ignored = []
    to_add_in_ignored.append(row[5])  # Date
    to_add_in_ignored.append(row[0])  # HC Id
    to_add_in_ignored.append(row[1])  # Node IP
    to_add_in_ignored.append(row[2])  # Location
    to_add_in_ignored.append(row[3])  # User Label
    # Handle missing HC IDs in ne_type_dict gracefully for ignored cases
    try:
        ne_type = ne_type_dict[int(row[0])]
    except KeyError:
        print(f"Warning: HC ID {int(row[0])} not found in NODE COVERAGE sheet, using 'Unknown'")
        ne_type = "Unknown"
    to_add_in_ignored.append(ne_type)  # NE Type - added 10th May 2024
    to_add_in_ignored.append(network_name)  # Network Name
    to_add_in_ignored.append(row[7])  # Test Case
    to_add_in_ignored.append(row[9])  # Category
    to_add_in_ignored.append(row[10])  # Issue
    to_add_in_ignored.append(row[11])  # Finding
    to_add_in_ignored.append(row[12])  # Task
    to_add_in_ignored.append("IGNORED")  # Status
    to_add_in_ignored.append("NA")  # Int/Ext
    to_add_in_ignored.append("NA")  # Fault Category
    to_add_in_ignored.append("NA")  # HW Type
    to_add_in_ignored.append("NA")  # Card Sl No.
    to_add_in_ignored.append("NA")  # Remarks
    master_sheet_ignored.append(to_add_in_ignored)

# 2a. Compare each OPEN entry in master tracker with extracted test cases (Only the W&F sheet in extracted cases is compared with the master tracker)
# First create an excel file "Closed in named [network_name + '_' + year_month + hc_report_date +xlsx] and activate, rename and save wbook
# Compare each OPEN entry in master tracker with extracted test cases
# If present do nothing (move on)
# If NOT present then it means that that issue was closed so CLOSE it in the master tracker
# Copy the entire just CLOSED row to closed in report file
# Copy the entire row to the HC tracker CLOSED sheet also
# Mark the case as CLOSED in the HC tracker OPEN sheet

# Create an excel file "Closed in named [network_name + '_' + year_month + hc_report_date +xlsx] and activate, rename and save wbook

closed_test_cases_report = opxl.Workbook()
closed_test_cases = closed_test_cases_report.active
closed_test_cases.title = "CLOSED in Master"

# Print the column titles in the closed test cases report
funcs.copy_first_row_hc_tracker(closed_test_cases, master_sheet_main)

# Filename for the closed test cases
closed_test_cases_filename = (
    "Closed in_" + network_name + "_" + year_month + hc_report_date + ".xlsx"
)

# Save deferred - will save at end for better performance
# closed_test_cases_report.save(output_dir / Path(closed_test_cases_filename))

# master_sheet details
# HC Id  	1
# Test Case	7
# Category  8
# Issue	    9
# Finding	10
# Status	11

# extracted wf details
# HCID	0
# Test	7
# Category 9
# Issue	10
# Findings	11

print("Checking for HC cases closed in this HC report...", end="")
# PERFORMANCE: Track time for large networks
step_start = time.time()

# Get total nodes for network size detection
total_nodes = (node_coverage_sheet.max_row) - 1

# Determine network size for processing optimization
if total_nodes < 100:
    detected_network_size = "SMALL"
elif total_nodes <= 1000:
    detected_network_size = "MEDIUM"
else:
    detected_network_size = "LARGE"

print(f"\n>> Network Size Detected: {detected_network_size} ({total_nodes} nodes)")

# REMOVE ALL TIMEOUTS FOR LARGE NETWORKS - NO SOCKET TIMEOUT LIMITS
if detected_network_size == "LARGE":
    # Remove any timeout limits to prevent failures
    socket.setdefaulttimeout(None)  # No timeout limit at all
    print(f"   [!] LARGE NETWORK PROCESSING - All timeout limits removed")
    print(f"   [*] Processing will take 30-60 minutes but will NOT fail")
    print(f"   [*] Memory optimizations enabled for large datasets")
    print(f"   [*] Progress will be shown during processing...")
    # Force garbage collection for large networks
    gc.collect()
else:
    # Keep reasonable timeout for smaller networks
    socket.setdefaulttimeout(1800)  # 30 minutes for medium/small networks
    print(f"   [T] Standard timeout: 30 minutes")

# OPTIMIZED: Build lookup set first for O(1) comparison instead of O(n²)
extracted_cases_set = set()
rows_processed = 0
for extracted_row in extracted_sheet_wf.iter_rows(min_row=2, values_only=True):
    if extracted_row and len(extracted_row) > 11:
        case_key = (extracted_row[0], extracted_row[7], extracted_row[9], extracted_row[10], extracted_row[11])
        extracted_cases_set.add(case_key)
        rows_processed += 1
        # Memory optimization for large networks
        if detected_network_size == "LARGE" and rows_processed % 5000 == 0:
            print(f" [Processing... {rows_processed} cases]...")
            gc.collect()  # Force garbage collection every 5000 rows

print(f" [Built lookup table with {len(extracted_cases_set)} extracted cases]...")

cases_closed_in_this_report = 0
closed_rows_to_process = []

# FAST: Single pass through master sheet with O(1) lookups
for row_index, m_row in enumerate(
    master_sheet_main.iter_rows(min_row=2, values_only=True), 2
):
    if m_row and len(m_row) > 12 and m_row[12] == "OPEN":
        master_case_key = (m_row[1], m_row[7], m_row[8], m_row[9], m_row[10])
        
        if master_case_key not in extracted_cases_set:
            # Case is closed - mark for processing
            master_sheet_main.cell(row=row_index, column=13).value = "CLOSED"
            closed_test_cases.append(m_row)
            closed_last_row = closed_test_cases.max_row
            closed_test_cases.cell(row=closed_last_row, column=13).value = "CLOSED"
            master_sheet_closed.append(m_row)
            master_sheet_closed_last_row = master_sheet_closed.max_row
            master_sheet_closed.cell(row=master_sheet_closed_last_row, column=13).value = "CLOSED"
            
            # Store for batch deletion from OPEN sheet
            closed_rows_to_process.append(m_row)
            cases_closed_in_this_report += 1

# OPTIMIZED: Batch delete from OPEN sheet
for m_row in closed_rows_to_process:
    funcs.delete_closed_case_open_sheet_hc_tracker(master_sheet_open, m_row)

# Save deferred - will save at end for better performance
# master_tracker.save(current_dir / Path(hc_issues_tracker_filename))
# Always save the closed cases report to ensure it appears in output directory
try:
    closed_test_cases_report.save(output_dir / Path(closed_test_cases_filename))
    print(f"Saved closed cases file: {closed_test_cases_filename} (cases_closed: {cases_closed_in_this_report})")
    # FORCE verify the file was actually saved
    saved_file_path = output_dir / Path(closed_test_cases_filename)
    if saved_file_path.exists():
        print(f"CONFIRMED: Closed cases file exists at {saved_file_path}")
    else:
        print(f"ERROR: Closed cases file was NOT saved properly!")
except Exception as e:
    print(f"Warning: Could not save closed cases file: {e}")
print("Done")

# Started working on this section from 13:17 22 March 2024 onwards

# Read-in the ignored tests for the network
ignored_tests = ignored_rows  # funcs.ignored_tests(current_dir, network_name)

# 2b. Compare EACH filtered extracted test case with the master tracker open cases
# open filtered extracted cases
# open master tracker
# create new cases workbook, rename sheet, set it active and save
# Compare EACH filtered extracted test case with the master tracker
# If present do nothing (move on) - case already reported and being worked on
# If not present then add this case to the master tracker for Int/Ext correction
# (How to decide Int/Ext? - for now use TBD)

# Create an excel file "New cases found in_[network_name + '_' + year_month + hc_report_date +xlsx] and activate, rename and save wbook

new_test_cases_found_report = opxl.Workbook()
new_test_cases_found = new_test_cases_found_report.active
new_test_cases_found.title = "NEW Cases found"

# Print the column titles in the new test cases found report
funcs.copy_first_row_hc_tracker(new_test_cases_found, cwbp_sheet)

# Filename for the new test cases found report
new_test_cases_found_filename = (
    "New cases found in_" + network_name + "_" + year_month + hc_report_date + ".xlsx"
)

# Save the new cases workbook as 'new test extracted_hc_test_cases.xlsx' in output_dir
new_test_cases_found_report.save(output_dir / Path(new_test_cases_found_filename))

print("Adding the new HC cases reported in this HC...", end="")
# OPTIMIZED: Build lookup set for existing OPEN cases in master tracker
master_open_cases_set = set()
master_rows_processed = 0
for m_row in master_sheet_main.iter_rows(min_row=2, values_only=True):
    if m_row and len(m_row) > 12 and m_row[12] == "OPEN":
        case_key = (m_row[1], m_row[7], m_row[8], m_row[9], m_row[10])
        master_open_cases_set.add(case_key)
        master_rows_processed += 1
        # Memory optimization for large networks
        if detected_network_size == "LARGE" and master_rows_processed % 2000 == 0:
            print(f" [Master processing... {master_rows_processed} cases]...")
            gc.collect()  # Force garbage collection every 2000 master rows

print(f" [Built master lookup table with {len(master_open_cases_set)} open cases]...")

new_cases_added_in_this_report = 0

# FAST: Single pass through filtered cases with O(1) lookups
batch_new_cases = []  # Batch processing for large networks
batch_size = 500 if detected_network_size == "LARGE" else 100

for ext_row in filtered_sheet.iter_rows(min_row=2, values_only=True):
    if ext_row and len(ext_row) > 11:
        ext_case_key = (int(ext_row[0]), ext_row[7], ext_row[9], ext_row[10], ext_row[11])
        
        if ext_case_key not in master_open_cases_set:
            # This is a new case - add it
            new_cases_list = [
                ext_row[5], ext_row[0], ext_row[1], ext_row[2], ext_row[3],
                ne_type_dict.get(ext_row[0], "Unknown"),
                network_name, ext_row[7], ext_row[9], ext_row[10], ext_row[11], ext_row[12],
                "OPEN", "Int", "TBD", " ", " ", " "
            ]
            
            # Batch processing for better performance
            batch_new_cases.append((new_cases_list, ext_row))
            new_cases_added_in_this_report += 1
            
            # Process batch when it reaches batch_size
            if len(batch_new_cases) >= batch_size:
                for new_case, ext_case in batch_new_cases:
                    master_sheet_main.append(new_case)
                    new_test_cases_found.append(ext_case)
                    master_sheet_open.append(new_case)
                
                batch_new_cases.clear()  # Clear batch
                if detected_network_size == "LARGE":
                    print(f" [Batch processed {new_cases_added_in_this_report} new cases]...")
                    gc.collect()  # Force garbage collection after each batch

# Process any remaining cases in the batch
if batch_new_cases:
    for new_case, ext_case in batch_new_cases:
        master_sheet_main.append(new_case)
        new_test_cases_found.append(ext_case)
        master_sheet_open.append(new_case)

print("Done")

# ADDED ON 12th April 2024
# Copy the Master tracker OPEN sheet to new workbook

# Generate the filename first
dt = datetime.today()  # gives us a datetime object
today_str = dt.strftime("%d-%b-%Y")  # format as 12-Apr-2024
tac_open_filename = network_name + "_OPEN cases for TAC_" + today_str + ".xlsx"

# Create the report workbook to store our test cases
tac_open_test_cases_report = opxl.Workbook()
tac_open_sheet = tac_open_test_cases_report.active
tac_open_sheet.title = "OPEN Cases for TAC"

# Copy the Master tracker OPEN sheet data to new workbook
for row in master_sheet_open.iter_rows(values_only=True):
    tac_open_sheet.append(row)

# Save the OPEN Cases for TAC workbook
tac_open_test_cases_report.save(output_dir / Path(tac_open_filename))

# Save deferred - will save at end for better performance
# master_tracker.save(current_dir / Path(hc_issues_tracker_filename))
# Save deferred - will save at end for better performance  
# new_test_cases_found_report.save(output_dir / Path(new_test_cases_found_filename))

# Define Cell comments

hc_run_comment = opxl.comments.Comment(
    text="Actual date when HC was run on the node", author="Automation Team"
)
hc_run_comment.width = 300
hc_run_comment.height = 30

hc_report_date_comment = opxl.comments.Comment(
    text="The heading date is the date when the TEC website created this report\nnode dates are the actual dates when HC was run on the nodes",
    author="Automation Team",
)
hc_report_date_comment.width = 300
hc_report_date_comment.height = 75

# Add cell comment to main tracker sheets 'MAIN', 'OPEN', 'CLOSED', 'IGNORED' in cell A1
for sheet in master_tracker.sheetnames:
    if master_tracker[sheet].cell(row=1, column=1).value == "HC Date":
        funcs.comment_cell(
            master_tracker[sheet].cell(row=1, column=1),
            hc_run_comment,
            comment_width=300,
            comment_height=30,
        )

# Node coverage will be updated after inventory is loaded

print("Formatting the HC issues tracker...", end="")

# Format the MAIN sheet of the HC issues tracker
funcs.format_worksheet(master_sheet_main)
# Format the CLOSED sheet of the HC issues tracker
funcs.format_worksheet(master_sheet_closed)
# Format the OPEN sheet of the HC issues tracker
funcs.format_worksheet(master_sheet_open)
# Format the Closed issues tracker
funcs.format_worksheet(closed_test_cases)
# Format the IGNORED sheet of the HC issues tracker
funcs.format_worksheet(master_sheet_ignored)
# Format the NODE COVERAGE sheet of the HC issues tracker
funcs.node_coverage_sheet_format(node_coverage_sheet)
# Format the the OPEN Cases for TAC workbook
funcs.format_worksheet(tac_open_sheet)

print("Done")

# print("Adding highlighting to OPEN cases found in this report...", end="")
# light_green = '00CCFFCC'
# sky_blue = '00CCECFF'
# white = '00FFFFFF'

# funcs.add_remove_highlight(master_sheet_main, filtered_sheet)
# funcs.add_remove_highlight(master_sheet_open, filtered_sheet)
# funcs.add_remove_highlight(tac_open_sheet, filtered_sheet)

# print("Done")

# Added on 11th Sep 2024 to remove highlighting
funcs.remove_highlight(master_sheet_main)
funcs.remove_highlight(master_sheet_open)
funcs.remove_highlight(tac_open_sheet)

# Added on 11th Sep 2024 to check for TC 1.1.1 and update NODE COVERAGE sheet
# Get the last column of NODE COVERAGE sheet for comments
last_row, last_column = funcs.get_last_row_col(node_coverage_sheet)

not_run_properly_comment = opxl.comments.Comment(
    text='HC has to run internally on node as appl environment found\nHC executed in wrong node environment.\nEnsure to have: \n- "expect" installed on your HC server\n- appl user (maint2) enabled by NECLI\n- protein.cfg correctly configured',
    author="Automation Team",
)
fill = opxl.styles.PatternFill(start_color="00FFC0CB", fill_type="solid")

appl_set = set()

for row in master_sheet_open.iter_rows(min_row=2, values_only=True):
    if (
        (row[7] == "1.1.1")
        and (row[10] == "HC has to run internally on node as appl environment found")
        and (row[12] == "OPEN")
    ):
        appl_set.add(row[1])

for row_number, row_data in enumerate(
    node_coverage_sheet.iter_rows(min_row=2, values_only=True), 2
):
    if row_data[0] in appl_set:
        funcs.comment_cell(
            node_coverage_sheet.cell(row=row_number, column=last_column),
            not_run_properly_comment,
            comment_width=300,
            comment_height=150,
        )
        node_coverage_sheet.cell(row=row_number, column=last_column).fill = fill

# Added on 19th Sep 2024 to check for TC 1.1.2 and update NODE COVERAGE sheet
not_run_properly_comment = opxl.comments.Comment(
    text="HC was stopped during NECLI session request\nIf this is related to a low performance of EC/FLC wait for a while and repeat HC run.\nIn all other cases issue internal Salesforce Case, assign it to TEC",
    author="Automation Team",
)
fill = opxl.styles.PatternFill(start_color="00CCECFF", fill_type="solid")

appl_set = set()

for row in master_sheet_open.iter_rows(min_row=2, values_only=True):
    if (
        (row[7] == "1.1.2")
        and (row[10] == "HC was stopped during NECLI session request")
        and (row[12] == "OPEN")
    ):
        appl_set.add(row[1])

for row_number, row_data in enumerate(
    node_coverage_sheet.iter_rows(min_row=2, values_only=True), 2
):
    if row_data[0] in appl_set:
        funcs.comment_cell(
            node_coverage_sheet.cell(row=row_number, column=last_column),
            not_run_properly_comment,
            comment_width=300,
            comment_height=100,
        )
        node_coverage_sheet.cell(row=row_number, column=last_column).fill = fill

# Added on 5th Nov 2024 to check for TC 22.0.5 and update NODE COVERAGE sheet
not_run_properly_comment = opxl.comments.Comment(
    text="Check Linux availability\nLinux session could not be open\nCheck if the used Linux user is enabled in NECLI and is configured in protein.cfg under pureONE line",
    author="Automation Team",
)
fill = opxl.styles.PatternFill(start_color="00FFFFBA", fill_type="solid")

appl_set = set()

for row in master_sheet_open.iter_rows(min_row=2, values_only=True):
    if (
        (row[7] == "22.0.5")
        and (row[10] == "Linux session could not be open")
        and (row[12] == "OPEN")
    ):
        appl_set.add(row[1])

for row_number, row_data in enumerate(
    node_coverage_sheet.iter_rows(min_row=2, values_only=True), 2
):
    if row_data[0] in appl_set:
        funcs.comment_cell(
            node_coverage_sheet.cell(row=row_number, column=last_column),
            not_run_properly_comment,
            comment_width=300,
            comment_height=100,
        )
        node_coverage_sheet.cell(row=row_number, column=last_column).fill = fill

print("Updating the Summary sheet...", end="")
# Added on 6th Nov 2024 for HC issues summary.
# Enter the HC report filename in Summary sheet cells F19 and A19
# F19 - For HC issues summary
master_sheet_summary["F19"].value = hc_filename
# A19 - For Node Coverage summary.
master_sheet_summary["A19"].value = hc_filename


# To find the total number of cases reported so far (OPEN & CLOSED together) is max row of sheet MAIN - 1
# B2 is Total number of cases (OPEN + CLOSED) in the MAIN sheet of the tracker
# master_sheet_summary['B2'].value = (master_sheet_main.max_row) - 1

# To find the total number of OPEN cases reported so far including the current report is max row of sheet OPEN - 1
# F20 is Total number of cases OPEN in the OPEN sheet of the tracker
master_sheet_summary["G20"].value = (master_sheet_open.max_row) - 1

# To find the total number of CLOSED cases reported so far including the current report is max row of sheet CLOSED - 1
# B4 is Total number of cases CLOSED in the CLOSED sheet of the tracker
# master_sheet_summary['B4'].value = (master_sheet_closed.max_row) - 1

# B3 is total number of new cases added in this report
master_sheet_summary["G21"].value = new_cases_added_in_this_report

# B4 is Total number of cases closed in this report
master_sheet_summary["G22"].value = cases_closed_in_this_report

# Added on 12th Nov 2024 for Node coverage summary.
# To find the Total No. of nodes in the network is max of node_coverage_sheet - 1
# B20 of the master_sheet_summary is the Total No. of nodes in the network
# Note: total_nodes already calculated earlier for network timeout configuration
master_sheet_summary["B20"].value = total_nodes

qty_nodes_covered, qty_nodes_not_covered, qty_nodes_not_run_properly = (
    funcs.summary_node_coverage(node_coverage_sheet, hc_report_month, hc_report_year)
)

# Nodes Covered in this HC report - if run dt mm-yyyy == report dt mm-yyyyy, then covered else not covered
# B21 of the master_sheet_summary is the total of Nodes Covered in this HC report
master_sheet_summary["B21"].value = qty_nodes_covered

# Nodes Not Covered in this HC report - if run dt mm-yyyy == report dt mm-yyyyy, then covered else not covered
# B22 of the master_sheet_summary is the total of Nodes Covered in this HC report
master_sheet_summary["B22"].value = qty_nodes_not_covered

# Nodes Not run properly in this HC report -  this is a subset of nodes covered this month
# B23 of the master_sheet_summary is the number of nodes where HC did not properly in this report
master_sheet_summary["B23"].value = qty_nodes_not_run_properly

print("Done")

# Added on 14th Nov 2024 for the remote inventory
print("Updating the Remote Inventory sheet...", end="")

# Get the sheet names of the master tracker
tracker_sheet_names = master_tracker.sheetnames

# Find the remote inventory sheet and make it active
for sheet_name in tracker_sheet_names:
    if (re.search(r'^\d{8} Remote Inventory', sheet_name)):
        rem_inv_sheet = master_tracker[sheet_name]
        break

# From row 2 of the remote inventory sheet, delete 10000 rows
rem_inv_sheet.delete_rows(idx=2, amount=10000)

# Rename the remote inventory sheet to the datestamp of the current report.
rem_inv_sheet.title = year_month_date + ' Remote Inventory'

# Set the path object for the remote inventory csv file
rem_inv_csv = input_hc_dir / Path(rem_inv_filename)

# Initialize an empty dict to hold the remote inventory data
rem_inv_dict = dict()

# Read the remote inventory from the CSV - skip the first row as it is the heading
rem_inv_list = []
with open(rem_inv_csv, 'r') as csv_input:
    csv_data = csv.reader(csv_input, delimiter=';')
    first_row = next(csv_data, None)
    # Add headers to the remote inventory sheet
    rem_inv_sheet.append(first_row)
    # Append to the remote inventory list to the rem_inv_list
    for row in csv_data:
       rem_inv_list.append(row)
       rem_inv_sheet.append(row)

# A28 - For Remote Inventory Summary.
master_sheet_summary["A28"].value = hc_filename            

# From row 29 of the Summary sheet, delete 100 rows
master_sheet_summary.delete_rows(idx=29, amount=100)

# Compile the remote inventory shelf wise
# Create a dictionary to store the count of mnemonics for each shelf type
mnemonic_count = {}

# Iterate over the rows in the sheet
for row in rem_inv_list:
    shelf_type = row[20]  # changed from 19 to 20 on 23rd June 2025
    mnemonic = row[14]  

    # If the shelf type is not in the dictionary, add it
    if shelf_type not in mnemonic_count:
        mnemonic_count[shelf_type] = {}

    # If the mnemonic is not in the dictionary for the shelf type, add it
    if mnemonic not in mnemonic_count[shelf_type]:
        mnemonic_count[shelf_type][mnemonic] = 1
    else:
        # If the mnemonic is already in the dictionary, increment the count
        mnemonic_count[shelf_type][mnemonic] += 1

# Get all unique mnemonics
all_mnemonics = set()
for shelf_type in mnemonic_count:
    all_mnemonics.update(mnemonic_count[shelf_type].keys())

# Create the header list
header = ['Board Mnemonic / Name'] + list(mnemonic_count.keys()) + ['Total']
if '' in all_mnemonics:
    all_mnemonics.remove('')
    all_mnemonics.add('Blank')

# Create the output list
output = [header]
for mnemonic in all_mnemonics:
    row = [mnemonic]
    total = 0
    for shelf_type in mnemonic_count:
        count = mnemonic_count[shelf_type].get(mnemonic, 0)
        row.append(count)
        total += count
    row.append(total)
    output.append(row)


for i, shelf_type in enumerate(output[0]):
    if shelf_type == '':
        output[0][i] = 'Blank' 

# Append the remote inventory from A29 onwards
# Start row and column
start_row = 29
start_col = 1  # Column A is column 1

# Write each row of data from the list into the worksheet
for i, row_data in enumerate(output):
    for j, value in enumerate(row_data):
        master_sheet_summary.cell(row=start_row + i, column=start_col + j).value = value

    
funcs.general_format_sheet(rem_inv_sheet)   # Format Rem Inv Sheet 20th Nov 2024
funcs.rem_inv_summary(master_sheet_summary)        # Format Rem Inv Sheet 20th Nov 2024

print("Done")

# NOW update the node coverage sheet after inventory is loaded
step_start = log_step_time("CSV_Processing", step_start)
print(f"Updating the node coverage sheet with inventory data ({len(rem_inv_list)} inventory rows)...", end="")
if len(rem_inv_list) == 0:
    print(f"\n[X] CRITICAL ISSUE: No inventory data loaded for {network_name}!")
    print(f"   - This causes NODE COVERAGE to show '1' in Location column")
    print(f"   - Remote Inventory sheet will have no proper headers")
    print(f"   - Make sure inventory CSV file is properly placed in input-hc-report directory!")
    print(f"   - Check CSV file permissions and format")
funcs.update_node_coverage(
    network_summary_sheet,
    node_coverage_sheet,
    hc_report_month,
    hc_report_year,
    hc_report_date,
)

# Find the last used column in NODE COVERAGE sheet and add the comment
node_coverage_sheet = master_tracker["NODE COVERAGE"]
last_row, last_column = funcs.get_last_row_col(node_coverage_sheet)
funcs.comment_cell(
    node_coverage_sheet.cell(row=1, column=last_column),
    hc_report_date_comment,
    comment_width=300,
    comment_height=75,
)
print("Done")

# Draw and save HC issues chart
funcs.draw_save_hc_issues_chart(master_sheet_summary, output_dir, hc_filename)

# Delete the old HC issues chart in the summary sheet
funcs.delete_hc_issues_chart(master_sheet_summary)

# Embed the HC issues chart in the summary sheet at F1
funcs.embed_chart(output_dir, hc_filename, master_sheet_summary)
#master_tracker.save(current_dir / Path(hc_issues_tracker_filename))
# Draw and save node coverage chart
funcs.draw_save_node_coverage_chart(master_sheet_summary, output_dir, hc_filename)

# Delete the old node coverage chart in the summary sheet
funcs.delete_node_coverage_chart(master_sheet_summary)

# Embed the Node coverage chart in the summary sheet at A1
funcs.embed_node_coverage_chart(output_dir, hc_filename, master_sheet_summary)
#master_tracker.save(current_dir / Path(hc_issues_tracker_filename))
# Save all files when the script completes
hc_test_cases_report.save(output_dir / Path(r"extracted_hc_test_cases.xlsx"))
# Ensure closed cases file is always saved, even if empty (for MOP compliance)
print(f"Final save - closed cases file rows: {closed_test_cases.max_row}, cases closed: {cases_closed_in_this_report}")

# FORCE save closed cases file with multiple attempts
saved_successfully = False
for attempt in range(3):
    try:
        closed_test_cases_report.save(output_dir / Path(closed_test_cases_filename))
        saved_file_path = output_dir / Path(closed_test_cases_filename)
        if saved_file_path.exists() and saved_file_path.stat().st_size > 0:
            print(f"FINAL SUCCESS: Closed cases file saved: {closed_test_cases_filename} (size: {saved_file_path.stat().st_size} bytes)")
            saved_successfully = True
            break
        else:
            print(f"Attempt {attempt + 1}: File save appeared successful but file missing or empty")
    except Exception as e:
        print(f"Attempt {attempt + 1} failed: {e}")
    
    if attempt < 2:  # Don't sleep on last attempt
        import time
        time.sleep(1)

if not saved_successfully:
    print(f"CRITICAL ERROR: Failed to save closed cases file after 3 attempts!")
    # Create a minimal closed cases file manually as fallback
    try:
        fallback_wb = opxl.Workbook()
        fallback_ws = fallback_wb.active
        fallback_ws.title = "CLOSED in Master"
        fallback_ws.append(["Date", "HC Id", "Node IP", "Location", "User Label", "NE Type", "Network Name", "Test Case", "Category", "Issue", "Finding", "Task", "Status", "Int/Ext", "Fault Category", "HW Type", "Card Sl No.", "Remarks"])
        fallback_wb.save(output_dir / Path(closed_test_cases_filename))
        print(f"FALLBACK SUCCESS: Created minimal closed cases file")
    except Exception as fallback_error:
        print(f"FALLBACK FAILED: {fallback_error}")
master_tracker.save(current_dir / Path(hc_issues_tracker_filename))
new_test_cases_found_report.save(output_dir / Path(new_test_cases_found_filename))
tac_open_test_cases_report.save(output_dir / Path(tac_open_filename))

# Added on 16th May 2024 on Niraj's request
last_hc_issues_tracker_filename = network_name + "_HC_Issues_Tracker"
this_report_date = "-".join((hc_report_date, hc_report_month, hc_report_year))
this_report_filename = (
    last_hc_issues_tracker_filename + "_" + this_report_date + ".xlsx"
)
master_tracker.save(output_dir / Path(this_report_filename))

# COMPREHENSIVE PERFORMANCE SUMMARY
script_end_time = time.time()
execution_time = script_end_time - script_start_time
step_start = log_step_time("Final_Operations", step_start)

print("\n" + "="*80)
print(f"*** PERFORMANCE OPTIMIZED PROCESSING COMPLETE FOR {network_name.upper()}! ***")
print("="*80)
print(f">> Total Execution Time: {execution_time:.2f}s ({execution_time/60:.2f} minutes)")

# Show performance analysis
if execution_time > 300:  # 5 minutes
    print(f"!!! PERFORMANCE WARNING: Processing took {execution_time/60:.1f} minutes")
    print(f"   - For large networks like Telekom, this may be normal")
    print(f"   - Consider running during off-peak hours for better performance")
elif execution_time > 180:  # 3 minutes
    print(f">>> ACCEPTABLE: Processing completed in {execution_time/60:.1f} minutes")
else:
    print(f">>> EXCELLENT: Fast processing completed in {execution_time:.1f} seconds")

print(f"\n>> Processing Results Summary:")
print(f"   - {len(filtered_data)} active test cases processed")
print(f"   - {new_cases_added_in_this_report} new cases added")
print(f"   - {cases_closed_in_this_report} cases closed")
print(f"   - {len(ignored_rows)} cases ignored")
print(f"\n>> Network Configuration:")
print(f"   - Network size: {detected_network_size} ({total_nodes} nodes)")
if detected_network_size == "LARGE":
    print(f"   - Large network processing mode enabled for optimal performance")

# Show critical issues if any
if len(rem_inv_list) == 0:
    print(f"\n!!! CRITICAL ISSUES DETECTED:")
    print(f"   [X] No inventory data loaded for {network_name}")
    print(f"   [X] NODE COVERAGE will show '1' in Location column")
    print(f"   [X] Remote Inventory sheet headers missing")
    print(f"   --> SOLUTION: Check inventory CSV file placement and format")
else:
    print(f"\n>>> All systems operational - {len(rem_inv_list)} inventory records loaded")

print(f"\n>> Output Files Generated:")
print(f"   - {hc_issues_tracker_filename} (Main tracker)")
print(f"   - {this_report_filename} (Date-stamped copy)")
print(f"   - {tac_open_filename} (TAC report)")
print(f"   - {closed_test_cases_filename} (Closed cases)")
print(f"   - {new_test_cases_found_filename} (New cases)")

print(f"\n>> Next Steps:")
print(f"   1. Open HC Issues Tracker: {hc_issues_tracker_filename}")
print(f"   2. Filter OPEN status to see active issues")
print(f"   3. Review NODE COVERAGE sheet for network coverage")
print(f"   4. Check Summary sheet for charts and statistics")
print("\n" + "="*80)

# Now close the workbook and exit
master_tracker.close()
sys.exit()
