# Utility functions used by 1830 HC script
import openpyxl as opxl
from datetime import datetime
from pathlib import Path
import matplotlib.pyplot as plt
import csv
import re

light_green = "00CCFFCC"
sky_blue = "00CCECFF"
white = "00FFFFFF"
rem_inv_blue = "0099CCFF"


def format_master_tracker(active_sheet):
    """Formats the master hc issues tracker

    Args:
    active_sheet (openpyxl worksheet object): the active sheet of the master tracker to be formatted
    """
    # Format the first row as follows
    # Font Calibri, Size 10, Bold Yes, H Alignment Center, V Alignment Center, Border Thin on all sides, Fill Solid Yellow

    font = opxl.styles.Font(name="Calibri", size=10, bold=True)
    alignment = opxl.styles.Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    thin_border = opxl.styles.Border(
        left=opxl.styles.Side(style="thin"),
        right=opxl.styles.Side(style="thin"),
        top=opxl.styles.Side(style="thin"),
        bottom=opxl.styles.Side(style="thin"),
    )
    fill = opxl.styles.PatternFill(start_color="00FFFF00", fill_type="solid")

    for i in range(1, active_sheet.max_column + 1):
        active_sheet.cell(row=1, column=i).font = font
        active_sheet.cell(row=1, column=i).alignment = alignment
        active_sheet.cell(row=1, column=i).fill = fill
        active_sheet.cell(row=1, column=i).border = thin_border

    # Format the second row and others below as follows
    # Font Calibri, Size 10, Bold No, H Alignment Center, V Alignment Center, Border Thin on all sides, No Fill

    font = opxl.styles.Font(name="Calibri", size=9, bold=False)

    for i in range(2, active_sheet.max_row + 1):
        for j in range(1, active_sheet.max_column + 1):
            active_sheet.cell(row=i, column=j).font = font
            active_sheet.cell(row=i, column=j).alignment = alignment
            active_sheet.cell(row=i, column=j).border = thin_border

    # Set the row height to 17 for the whole sheet
    for row in range(1, active_sheet.max_row + 1):
        active_sheet.row_dimensions[row].height = 17

    # Set the column width of column A(Date) to 9
    active_sheet.column_dimensions["A"].width = 9

    # Set the column width of column B(HC Id) to 7
    active_sheet.column_dimensions["B"].width = 7

    # Set the column width of each column to max length in column value + 5

    for col in range(3, active_sheet.max_column + 1):
        column_letter = opxl.utils.cell.get_column_letter(
            col
        )  # Get col letter from col number
        max_col_width = 0
        for row in range(1, active_sheet.max_row + 1):
            column_length = len(active_sheet.cell(row=row, column=col).value)
            # column_width = active_sheet.column_dimensions[column_letter].width  # Get the col width for the current row
            if column_length > max_col_width:
                max_col_width = column_length
        active_sheet.column_dimensions[column_letter].width = max_col_width + 5

    # Added from here on 28th March 2024
    # First find the row number from which the new cases were added by looking type str
    for row_number, date_value in enumerate(
        active_sheet.iter_rows(min_row=2, values_only=True), 2
    ):
        if isinstance(date_value[0], str):
            break

    # Format first column as date object and second column as int from that row till end of sheet
    for row in range(row_number, active_sheet.max_row + 1):
        # Format date first
        date_string = active_sheet.cell(row=row, column=1).value
        # We want this as a date in the format dd-mmm-yy
        # date_obj = datetime.strptime(date_string, "%Y%m%d")
        # formatted_date = date_obj.strftime("%d-%b-%y")
        year = date_string[:4]
        month = date_string[4:6]
        date = date_string[6:]
        active_sheet.cell(row=row, column=1).value = datetime(
            int(year), int(month), int(date)
        )
        active_sheet.cell(row=row, column=1).number_format = "d-mmm-yy"
        # Format the HC Id column as an int
        active_sheet.cell(row=row, column=2).value = int(
            active_sheet.cell(row=row, column=2).value
        )

    # my_book.save(format_test_tracker)
    return


def copy_first_row_hc_tracker(active_sheet, sheet_to_copy_from):
    for i in range(1, sheet_to_copy_from.max_column + 1):
        active_sheet.cell(row=1, column=i).value = sheet_to_copy_from.cell(
            row=1, column=i
        ).value


def delete_closed_case_open_sheet_hc_tracker(open_sheet, m_row):
    for o_row in reversed(list(open_sheet.iter_rows())):
        # Find the row to be deleted
        if (
            (o_row[0].value == m_row[0])
            and (o_row[1].value == m_row[1])
            and (o_row[2].value == m_row[2])
            and (o_row[3].value == m_row[3])
            and (o_row[4].value == m_row[4])
            and (o_row[5].value == m_row[5])
            and (o_row[6].value == m_row[6])
            and (o_row[7].value == m_row[7])
            and (o_row[8].value == m_row[8])
            and (o_row[9].value == m_row[9])
            and (o_row[10].value == m_row[10])  # added on 1st Oct'24
            and (o_row[11].value == m_row[11])
            and (o_row[12].value == m_row[12])
            and (o_row[13].value == m_row[13])
            and (o_row[14].value == m_row[14])
            and (o_row[15].value == m_row[15])
        ):
            # to_delete = True
            # Delete the entire row
            open_sheet.delete_rows(o_row[0].row, 1)
            # import pdb; pdb.set_trace()


def general_format_sheet(active_sheet):
    # Format the first row as follows
    # Font Calibri, Size 10, Bold Yes, H Alignment Center, V Alignment Center, Border Thin on all sides, Fill Solid Yellow

    font = opxl.styles.Font(name="Calibri", size=10, bold=True)
    alignment = opxl.styles.Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    thin_border = opxl.styles.Border(
        left=opxl.styles.Side(style="thin"),
        right=opxl.styles.Side(style="thin"),
        top=opxl.styles.Side(style="thin"),
        bottom=opxl.styles.Side(style="thin"),
    )
    fill = opxl.styles.PatternFill(start_color="00FFFF00", fill_type="solid")

    for i in range(1, active_sheet.max_column + 1):
        active_sheet.cell(row=1, column=i).font = font
        active_sheet.cell(row=1, column=i).alignment = alignment
        active_sheet.cell(row=1, column=i).fill = fill
        active_sheet.cell(row=1, column=i).border = thin_border

    # Format the second row and others below as follows
    # Font Calibri, Size 10, Bold No, H Alignment Center, V Alignment Center, Border Thin on all sides, No Fill

    font = opxl.styles.Font(name="Calibri", size=9, bold=False)

    for i in range(2, active_sheet.max_row + 1):
        for j in range(1, active_sheet.max_column + 1):
            active_sheet.cell(row=i, column=j).font = font
            active_sheet.cell(row=i, column=j).alignment = alignment
            active_sheet.cell(row=i, column=j).border = thin_border

    # Set the row height to 17 for the whole sheet
    for row in range(1, active_sheet.max_row + 1):
        active_sheet.row_dimensions[row].height = 17

    return


def format_date_column(active_sheet, column_number):
    # Check if this is a date column
    sub_string = "date"
    heading = (active_sheet.cell(row=1, column=column_number).value).lower()
    if sub_string in heading:
        # Get column letter
        column_letter = opxl.utils.cell.get_column_letter(column_number)
        # Set the column width of the date column to 9
        active_sheet.column_dimensions[column_letter].width = 9
        for row_number in range(2, active_sheet.max_row + 1):
            date_string = active_sheet.cell(row=row_number, column=column_number).value
            if isinstance(date_string, datetime):
                active_sheet.cell(
                    row=row_number, column=column_number
                ).number_format = "d-mmm-yy"
            elif isinstance(date_string, str):
                year = date_string[:4]
                month = date_string[4:6]
                date = date_string[6:]
                active_sheet.cell(
                    row=row_number, column=column_number
                ).value = datetime(int(year), int(month), int(date))
                active_sheet.cell(row=row_number, column=1).number_format = "d-mmm-yy"
    return


def format_hc_id_column(active_sheet, column_number):
    # HC Id is an integer
    # Check if this is a hc id column
    sub_string = "hc id"
    heading = (active_sheet.cell(row=1, column=column_number).value).lower()
    if sub_string in heading:
        # Get column letter
        column_letter = opxl.utils.cell.get_column_letter(column_number)
        # Set the column width of the date column to 7
        active_sheet.column_dimensions[column_letter].width = 7
        for row_number in range(2, active_sheet.max_row + 1):
            cell_value = active_sheet.cell(row=row_number, column=column_number).value
            if isinstance(cell_value, str):
                active_sheet.cell(row=row_number, column=column_number).value = int(
                    cell_value
                )
    return


def format_text_column(active_sheet, column_number):
    # Format this column as a text column only if this col is neither a date nor a number (hc id) col
    date_sub_string = "date"
    integer_sub_string = "hc id"
    heading = (active_sheet.cell(row=1, column=column_number).value).lower()
    if (date_sub_string not in heading) and (integer_sub_string not in heading):
        # Get column letter
        column_letter = opxl.utils.cell.get_column_letter(column_number)
        # Set the column width to max length in column value + 5
        max_col_width = 0
        for row in range(1, active_sheet.max_row + 1):
            # Added on 13th Jan 2025
            if active_sheet.cell(row=row, column=column_number).value is not None:
                column_length = len(active_sheet.cell(row=row, column=column_number).value)
                if column_length > max_col_width:
                    max_col_width = column_length
        active_sheet.column_dimensions[column_letter].width = max_col_width + 5
    return


def format_worksheet(active_sheet):
    for col in range(1, active_sheet.max_column + 1):
        format_date_column(active_sheet, col)
        format_hc_id_column(active_sheet, col)
        format_text_column(active_sheet, col)
    general_format_sheet(active_sheet)
    return


def ignored_tests(working_dir, network_name):
    # Get the filename of the ignored test cases for this network
    ignored_filename = network_name + "_ignored_test_cases.txt"
    if Path.exists(working_dir / Path(ignored_filename)):
        # Read-in the ignored test cases for this network from file
        ignored_cases = working_dir / Path(ignored_filename)
        with open(ignored_cases) as f:
            ignored_test_cases = f.readlines()
        # Strip all combinations of line endings - all combinations of \r and \n
        ignored_test_cases = [line.rstrip("\r\n") for line in ignored_test_cases]
        # Delete blank line at the end if any
        if ignored_test_cases[-1] == "":
            ignored_test_cases.pop()
        return ignored_test_cases
    else:
        ignored_test_cases = []
        return ignored_test_cases


def update_node_coverage(
    network_summary_sheet,
    node_coverage_sheet,
    hc_report_month,
    hc_report_year,
    hc_report_date,
):
    # Get the last used row number and column_number of both sheets
    summary_sheet_last_row, summary_sheet_last_col = get_last_row_col(
        network_summary_sheet
    )
    nc_last_row, nc_last_col = get_last_row_col(node_coverage_sheet)
    new_col_idx = nc_last_col + 1
    node_coverage_sheet.insert_cols(idx=new_col_idx)
    # Give a heading to the newly inserted col
    # Set the header value and format
    header_cell = node_coverage_sheet.cell(row=1, column=new_col_idx)
    header_cell.value = datetime(
        int(hc_report_year), int(hc_report_month), int(hc_report_date)
    )
    header_cell.number_format = "dd-mmm-yy"
    
    # Apply header formatting to match other columns
    import openpyxl as opxl
    font = opxl.styles.Font(name="Calibri", size=10, bold=True)
    alignment = opxl.styles.Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    thin_border = opxl.styles.Border(
        left=opxl.styles.Side(style="thin"),
        right=opxl.styles.Side(style="thin"),
        top=opxl.styles.Side(style="thin"),
        bottom=opxl.styles.Side(style="thin"),
    )
    fill = opxl.styles.PatternFill(start_color="00FFFF00", fill_type="solid")
    
    header_cell.font = font
    header_cell.alignment = alignment
    header_cell.fill = fill
    header_cell.border = thin_border
    
    # Set column width to match other date columns
    column_letter = opxl.utils.cell.get_column_letter(new_col_idx)
    node_coverage_sheet.column_dimensions[column_letter].width = 12

    for row_number, t_row in enumerate(
        node_coverage_sheet.iter_rows(min_row=2, max_row=nc_last_row, values_only=True),
        2,
    ):
        tracker_hc_id = t_row[0]
        found = False
        for h_row in network_summary_sheet.iter_rows(
            min_row=2, max_row=summary_sheet_last_row, values_only=True
        ):
            tec_hc_id = int(h_row[1])
            if tracker_hc_id == tec_hc_id:
                found = True
                hc_run_date = h_row[2]
                data_cell = node_coverage_sheet.cell(
                    row=row_number, column=new_col_idx
                )
                data_cell.value = hc_run_date
                
                # Apply formatting to data cells
                data_font = opxl.styles.Font(name="Calibri", size=9, bold=False)
                data_alignment = opxl.styles.Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                data_cell.font = data_font
                data_cell.alignment = data_alignment
                data_cell.border = thin_border
                
                if isinstance(hc_run_date, datetime):
                    data_cell.number_format = "dd-mmm-yy"
                break
        if not found:
            missing_cell = node_coverage_sheet.cell(
                row=row_number, column=new_col_idx
            )
            missing_cell.value = "Missing"
            
            # Apply formatting to missing cells too
            data_font = opxl.styles.Font(name="Calibri", size=9, bold=False)
            data_alignment = opxl.styles.Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            missing_cell.font = data_font
            missing_cell.alignment = data_alignment
            missing_cell.border = thin_border
    return


def get_last_row_col(active_sheet):
    # Find the last row with data
    last_row = 0
    for row_number, row in enumerate(active_sheet.iter_rows(values_only=True), 1):
        if row[0] is not None:
            last_row += 1
        else:
            break

    # Find the last column with data
    last_column = 0
    for col_number, col in enumerate(active_sheet.iter_cols(values_only=True), 1):
        if col[0] is not None:
            last_column += 1
        else:
            break

    return (last_row, last_column)


def node_coverage_sheet_format(active_sheet):
    # First apply general formatting
    general_format_sheet(active_sheet)
    
    # Then apply specific column formatting
    for col in range(1, active_sheet.max_column + 1):
        format_nc_date_column(active_sheet, col)
        format_nc_hc_id_column(active_sheet, col)
        format_nc_text_column(active_sheet, col)
    return


def format_nc_date_column(active_sheet, column_number):
    # Check if this is a date column
    heading = active_sheet.cell(row=1, column=column_number).value
    if isinstance(heading, datetime):
        # Get column letter
        column_letter = opxl.utils.cell.get_column_letter(column_number)
        # Set the column width of the date column to 12
        active_sheet.column_dimensions[column_letter].width = 12
        for row_number in range(2, active_sheet.max_row + 1):
            date_string = active_sheet.cell(row=row_number, column=column_number).value
            if isinstance(date_string, datetime):
                active_sheet.cell(
                    row=row_number, column=column_number
                ).number_format = "dd-mmm-yy"
            elif isinstance(date_string, str):
                continue
    return


def format_nc_text_column(active_sheet, col):
    # Format this column as a text column only if this col is neither a date nor a number (hc id) col
    heading = active_sheet.cell(row=1, column=col).value
    if not isinstance(heading, datetime) and heading != "HC Id":
        # Get column letter
        column_letter = opxl.utils.cell.get_column_letter(col)
        # Set the column width to max length in column value + 5
        max_col_width = 0
        for row in range(1, active_sheet.max_row + 1):
            column_length = len(active_sheet.cell(row=row, column=col).value)
            if column_length > max_col_width:
                max_col_width = column_length
        active_sheet.column_dimensions[column_letter].width = max_col_width + 5
    return


def format_nc_hc_id_column(active_sheet, col):
    if active_sheet.cell(row=1, column=col).value == "HC Id":
        # Get column letter
        column_letter = opxl.utils.cell.get_column_letter(col)
        # Set the column width of the date column to 7
        active_sheet.column_dimensions[column_letter].width = 7
    return


def extracted_sheet_format_hc_id_column(active_sheet, column_number):
    # HC Id is an integer
    # Check if this is a hc id column
    sub_string = "hcid"
    heading = (active_sheet.cell(row=1, column=column_number).value).lower()
    if sub_string in heading:
        # Get column letter
        column_letter = opxl.utils.cell.get_column_letter(column_number)
        # Set the column width of the date column to 7
        active_sheet.column_dimensions[column_letter].width = 7
        for row_number in range(2, active_sheet.max_row + 1):
            cell_value = active_sheet.cell(row=row_number, column=column_number).value
            if isinstance(cell_value, str):
                active_sheet.cell(row=row_number, column=column_number).value = int(
                    cell_value
                )
    return


def remove_ignore_from_extracted(extracted_sheet, ignored_sheet, indices_to_check):
    # From extracted test cases sheet extract a list of tuples - each row is a tuple
    extracted_list = [
        row for row in extracted_sheet.iter_rows(min_row=2, values_only=True)
    ]
    # From ignored test cases sheet extract a list of tuples - each row is a tuple
    ignored_list = [row for row in ignored_sheet.iter_rows(min_row=2, values_only=True)]

    extracted_set = {tuple(row[i] for i in indices_to_check) for row in extracted_list}
    ignored_set = {tuple(row[i] for i in indices_to_check) for row in ignored_list}

    filtered_indices = extracted_set.difference(ignored_set)

    filtered_data = []
    ignored_rows = []
    for i in range(len(extracted_list)):
        if tuple(extracted_list[i][j] for j in indices_to_check) in filtered_indices:
            filtered_data.append(extracted_list[i])
        else:
            ignored_rows.append(extracted_list[i])

    return (filtered_data, ignored_rows)


def comment_cell(cell_to_comment, the_comment, comment_width, comment_height):
    cell_to_comment.comment = the_comment
    the_comment.width = comment_width
    the_comment.height = comment_height
    return


def remove_highlight(active_sheet):
    # Remove all cell highlighting in open_sheet
    fill = opxl.styles.PatternFill(start_color="00FFFFFF", fill_type="solid")

    for row in range(2, active_sheet.max_row + 1):
        for col in range(1, active_sheet.max_column + 1):
            active_sheet.cell(row=row, column=col).fill = fill
    return


def add_highlight(active_sheet, start_row, fill_color):
    # Highlight all cells in open_sheet
    fill = opxl.styles.PatternFill(start_color=fill_color, fill_type="solid")

    for row in range(start_row, active_sheet.max_row + 1):
        for col in range(1, active_sheet.max_column + 1):
            active_sheet.cell(row=row, column=col).fill = fill
    return


def get_earliest_date(filtered_sheet):
    # Get the earliest date in the 'filtered_from_extracted_hc_test_cases.xlsx' worksheet
    min_date_string = filtered_sheet.cell(row=2, column=6).value
    min_date = datetime.strptime(min_date_string, "%Y%m%d")

    for row in filtered_sheet.iter_rows(min_row=2, values_only=True):
        if datetime.strptime(row[5], "%Y%m%d") < min_date:
            min_date = datetime.strptime(row[5], "%Y%m%d")
    return min_date


def get_row_number_of_earliest_date(active_sheet, min_date):
    for row_number, row in enumerate(
        active_sheet.iter_rows(min_row=2, values_only=True), 2
    ):
        if (row[0] == min_date) and (row[11] == "OPEN"):
            break
    return row_number


def add_remove_highlight(active_sheet, filtered_sheet):
    earliest_date = get_earliest_date(filtered_sheet)
    start_row = get_row_number_of_earliest_date(active_sheet, earliest_date)
    remove_highlight(active_sheet)
    add_highlight(active_sheet, start_row, fill_color=sky_blue)
    return


# Added on 12th Nov 2024 to calculate summary of node coverage
def summary_node_coverage(node_coverage_sheet, hc_report_month, hc_report_year):
    total_nodes_covered_in_this_report = 0
    total_nodes_not_covered_in_this_report = 0
    total_nodes_not_run_properly = 0

    for i in range(2, (node_coverage_sheet.max_row) + 1):
        node_hc_run_date = node_coverage_sheet.cell(
            row=i, column=node_coverage_sheet.max_column
        ).value
        if isinstance(node_hc_run_date, str):
            total_nodes_not_covered_in_this_report += 1
            continue
        else:
            node_hc_run_month = datetime.strftime(node_hc_run_date, "%m")
            node_hc_run_year = datetime.strftime(node_hc_run_date, "%Y")

            if (node_hc_run_month == hc_report_month) and (
                node_hc_run_year == hc_report_year
            ):
                total_nodes_covered_in_this_report += 1
            else:
                total_nodes_not_covered_in_this_report += 1

            if (
                node_coverage_sheet.cell(
                    row=i, column=node_coverage_sheet.max_column
                ).comment
            ) is not None:
                total_nodes_not_run_properly += 1

    return (
        total_nodes_covered_in_this_report,
        total_nodes_not_covered_in_this_report,
        total_nodes_not_run_properly,
    )


# Added on 20th Nov 2024 for formatting the remote inventory summary
def rem_inv_summary(active_sheet):
    # Font Calibri, Size 9, Bold No, H Alignment Center, V Alignment Center, Border Thin on all sides, No Fill
    font = opxl.styles.Font(name="Calibri", size=9, bold=False)
    alignment = opxl.styles.Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    thin_border = opxl.styles.Border(
        left=opxl.styles.Side(style="thin"),
        right=opxl.styles.Side(style="thin"),
        top=opxl.styles.Side(style="thin"),
        bottom=opxl.styles.Side(style="thin"),
    )

    # Added on 6th Jan 2025 to format the first line of the Rem Inv summary
    # Find how many columns are there for Rem Inv summary
    rem_inv_summary_max_col = 0
    for i in range(1, 20):
        if active_sheet.cell(row=29, column=i).value is not None:
            rem_inv_summary_max_col += 1
        else:
            break
  

    for i in range(30, active_sheet.max_row + 1):
        for j in range(1, rem_inv_summary_max_col +1):
            active_sheet.cell(row=i, column=j).font = font
            active_sheet.cell(row=i, column=j).alignment = alignment
            active_sheet.cell(row=i, column=j).border = thin_border

    # Set the row height to 17 for the whole sheet
    for row in range(30, active_sheet.max_row + 1):
        active_sheet.row_dimensions[row].height = 17

    # Format the rem inv summary row
    # Font Calibri, Size 10, Bold Yes, H Alignment Center, V Alignment Center, Border Thin on all sides, Fill rem_inv_blue

    font = opxl.styles.Font(name="Calibri", size=10, bold=True)
    alignment = opxl.styles.Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    thin_border = opxl.styles.Border(
        left=opxl.styles.Side(style="thin"),
        right=opxl.styles.Side(style="thin"),
        top=opxl.styles.Side(style="thin"),
        bottom=opxl.styles.Side(style="thin"),
    )
    fill = opxl.styles.PatternFill(start_color="0099CCFF", fill_type="solid")

    for i in range(1, rem_inv_summary_max_col + 1):
        active_sheet.cell(row=29, column=i).font = font
        active_sheet.cell(row=29, column=i).alignment = alignment
        active_sheet.cell(row=29, column=i).fill = fill
        active_sheet.cell(row=29, column=i).border = thin_border

    return


def draw_save_hc_issues_chart(master_sheet_summary, output_dir, hc_filename):
    # Data for the chart
    categories = [
        "TOTAL OPEN CASES AFTER ANALYZING THIS HC REPORT",
        "NEW CASES ADDED IN THIS HC REPORT",
        "CASES CLOSED IN THIS HC REPORT",
    ]
    values = [
        master_sheet_summary["G20"].value,
        master_sheet_summary["G21"].value,
        master_sheet_summary["G22"].value,
    ]
    colors = ["#50c4d0", "#f7b737", "#37cc73"]
    background_color = "#badaff"  # Light blue color

    # Create the horizontal bar chart with slightly more height
    fig, ax = plt.subplots(
        figsize=(10, 6.5)
    )  # Slightly increased height to accommodate legend
    fig.patch.set_facecolor(background_color)
    ax.set_facecolor(background_color)

    # Plot the bars
    bars = ax.barh(categories, values, color=colors)

    # Add value labels at the end of each bar
    for bar, value in zip(bars, values):
        plt.text(
            bar.get_width() + 1,
            bar.get_y() + bar.get_height() / 2,
            f"{value}",
            ha="left",
            va="center",
            fontsize=10,
            color="black",
            fontweight="bold",
        )

    # Add title and labels
    plt.title("Summary of HC Issues - " + hc_filename, fontsize=14, fontweight="bold")
    plt.xlabel("")
    plt.ylabel("")

    # Add some padding to x-axis to accommodate labels
    ax.set_xlim(right=max(values) * 1.1)

    # First adjust the subplot parameters
    plt.subplots_adjust(bottom=0.2)  # Set this before tight_layout

    # Add legend with adjusted positioning
    plt.legend(
        bars,
        [
            "Total OPEN cases after analyzing this HC report",
            "New Cases added in this HC Report",
            "Cases CLOSED in this HC report",
        ],
        loc="upper center",
        bbox_to_anchor=(0.5, -0.08),
        ncol=3,
        fontsize=10,
        frameon=True,
    )

    # Remove x-axis grid
    plt.grid(False)

    # Save the plot
    chart_filename = hc_filename[: (len(hc_filename) - 4)]
    plt.savefig(
        output_dir / Path(chart_filename + ".png"),
        bbox_inches="tight",
        dpi=300,
        pad_inches=0.5,
    )

    return


def delete_hc_issues_chart(active_sheet):
    # Remove the image at F1
    # Note: This assumes there's an image anchored at cell F1
    for image in active_sheet._images[:]:
        if image.anchor._from.col == 5 and image.anchor._from.row == 0:
            active_sheet._images.remove(image)
            del image

        


def embed_chart(output_dir, hc_filename, active_sheet):
    # Point to the image
    chart_filename = hc_filename[: (len(hc_filename) - 4)]
    img = opxl.drawing.image.Image(output_dir / Path(chart_filename + ".png"))

    # Position the image in cell F1 and adjust its size
    img.anchor = "F1"
    img.width = 600
    img.height = 300

    # Add the image to the worksheet
    active_sheet.add_image(img)

    return


def draw_save_node_coverage_chart(master_sheet_summary, output_dir, hc_filename):
    # Data for the chart
    categories = [
        "TOTAL NO. OF NODES IN THE NETWORK",
        "NODES COVERED IN THIS HC REPORT",
        "NODES NOT COVERED IN THIS HC REPORT",
        "NODES WHERE HC DID NOT RUN PROPERLY IN THIS HC REPORT",
    ]
    values = [
        master_sheet_summary["B20"].value,
        master_sheet_summary["B21"].value,
        master_sheet_summary["B22"].value,
        master_sheet_summary["B23"].value,
    ]
    colors = ["#50c4d0", "#37cc73", "#f7b737", "#ff6b6b"]
    background_color = "#badaff"  # Light blue color

    # Create the horizontal bar chart with slightly more height
    fig, ax = plt.subplots(figsize=(10, 6.5))
    fig.patch.set_facecolor(background_color)
    ax.set_facecolor(background_color)

    # Plot the bars
    bars = ax.barh(categories, values, color=colors)

    # Add value labels at the end of each bar
    for bar, value in zip(bars, values):
        plt.text(
            bar.get_width() + 1,
            bar.get_y() + bar.get_height() / 2,
            f"{value}",
            ha="left",
            va="center",
            fontsize=10,
            color="black",
            fontweight="bold",
        )

    # Add title and labels
    plt.title("Node Coverage Summary - " + hc_filename , fontsize=14, fontweight="bold")
    plt.xlabel("")
    plt.ylabel("")

    # Add some padding to x-axis to accommodate labels
    ax.set_xlim(right=max(values) * 1.1)

    # First adjust the subplot parameters
    plt.subplots_adjust(bottom=0.2)

    # Add legend with adjusted positioning
    plt.legend(
        bars,
        [
            "Total no. of nodes in the network",
            "Nodes Covered in this HC report",
            "Nodes Not Covered in this HC report",
            "Nodes where HC did not run properly in this HC report",
        ],
        loc="upper center",
        bbox_to_anchor=(0.5, -0.08),
        ncol=2,  # Changed to 2 columns for better readability of longer text
        fontsize=10,
        frameon=True,
    )

    # Remove x-axis grid
    plt.grid(False)

    # Save the plot
    chart_filename = hc_filename[: (len(hc_filename) - 4)]
    plt.savefig(
        output_dir / Path('Node Coverage ' + chart_filename + ".png"),
        bbox_inches="tight",
        dpi=300,
        pad_inches=0.5,
    )

    return

# def delete_node_coverage_chart(active_sheet): # my function 20/12/24
#     # Remove the image at A1
#     # Note: This assumes there's an image anchored at cell A1
#     for image in active_sheet._images[:]:
#         if image.anchor._from.col == 0 and image.anchor._from.row == 0:
#             active_sheet._images.remove(image)


def delete_node_coverage_chart(active_sheet):
    # Remove the image at A1
    if hasattr(active_sheet, '_images'):
        for image in active_sheet._images[:]:
            try:
                if (hasattr(image, 'anchor') and 
                    hasattr(image.anchor, '_from') and 
                    hasattr(image.anchor._from, 'col') and 
                    hasattr(image.anchor._from, 'row')):
                    
                    if image.anchor._from.col == 0 and image.anchor._from.row == 0:
                        active_sheet._images.remove(image)
                        del image
                        
            except AttributeError:
                # Skip images that don't have the expected attributes
                continue

    return


def format_summary_sheet(active_sheet):
    """Format Summary sheet with proper boxes and formatting like HC Issues Tracker"""
    
    # SECTION 1: HC Issues Summary (Rows 19-25)
    # Add borders and formatting for HC Issues section
    hc_summary_range = [(19, 6, 22, 8)]  # F19:H22
    
    # SECTION 2: Node Coverage Summary (Rows 19-25) 
    # Add borders and formatting for Node Coverage section
    node_summary_range = [(19, 1, 25, 3)]  # A19:C25
    
    # SECTION 3: Remote Inventory Summary (Row 28 onwards)
    # Already formatted by rem_inv_summary function
    
    # Apply formatting to HC Issues summary section
    for start_row, start_col, end_row, end_col in hc_summary_range:
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = active_sheet.cell(row=row, column=col)
                
                # Add borders
                cell.border = opxl.styles.Border(
                    left=opxl.styles.Side(style="thin"),
                    right=opxl.styles.Side(style="thin"),
                    top=opxl.styles.Side(style="thin"),
                    bottom=opxl.styles.Side(style="thin")
                )
                
                # Header row formatting (F19, G19, H19)
                if row == 19:
                    cell.font = opxl.styles.Font(name="Calibri", size=10, bold=True)
                    cell.fill = opxl.styles.PatternFill(start_color="00FFFF00", fill_type="solid")
                    cell.alignment = opxl.styles.Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )
                else:
                    # Data rows formatting
                    cell.font = opxl.styles.Font(name="Calibri", size=9, bold=False)
                    cell.alignment = opxl.styles.Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )
    
    # Apply formatting to Node Coverage summary section
    for start_row, start_col, end_row, end_col in node_summary_range:
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = active_sheet.cell(row=row, column=col)
                
                # Add borders
                cell.border = opxl.styles.Border(
                    left=opxl.styles.Side(style="thin"),
                    right=opxl.styles.Side(style="thin"),
                    top=opxl.styles.Side(style="thin"),
                    bottom=opxl.styles.Side(style="thin")
                )
                
                # Header row formatting (A19, B19, C19)
                if row == 19:
                    cell.font = opxl.styles.Font(name="Calibri", size=10, bold=True)
                    cell.fill = opxl.styles.PatternFill(start_color="00FFFF00", fill_type="solid")
                    cell.alignment = opxl.styles.Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )
                else:
                    # Data rows formatting
                    cell.font = opxl.styles.Font(name="Calibri", size=9, bold=False)
                    cell.alignment = opxl.styles.Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )
    
    # Set appropriate column widths
    active_sheet.column_dimensions['A'].width = 15
    active_sheet.column_dimensions['B'].width = 12
    active_sheet.column_dimensions['C'].width = 12
    active_sheet.column_dimensions['F'].width = 15
    active_sheet.column_dimensions['G'].width = 12
    active_sheet.column_dimensions['H'].width = 12
    
    # Set row heights
    for row_num in range(19, 26):
        active_sheet.row_dimensions[row_num].height = 17
    
    return

def embed_node_coverage_chart(output_dir, hc_filename, active_sheet):
    # Point to the image
    chart_filename = hc_filename[: (len(hc_filename) - 4)]
    img = opxl.drawing.image.Image(output_dir / Path('Node Coverage ' + chart_filename + ".png"))

    # Position the image in cell A1 and adjust its size
    img.anchor = "A1"
    img.width = 600
    img.height = 300

    # Add the image to the worksheet
    active_sheet.add_image(img)

    return

def update_pss_type(network_summary_sheet, node_coverage_sheet):
    # Determine the last entry in network summary sheet - value Total Network Issues - 2
    last_node_row = 0
    for row in network_summary_sheet.iter_rows(min_row=1, max_row=network_summary_sheet.max_row, values_only=True):
        if row[0] != 'Total Network Issues':
            last_node_row += 1
        else:
            last_node_row = last_node_row - 1
            break

    # Read in HC Id ,Node Type ,Location ,SID (System ID) and PSS Type of network summary sheet
    nw_summ_sh_data_dict = dict()
    for row in network_summary_sheet.iter_rows(min_row=2, max_row = last_node_row,values_only=True):
        hcid = int(row[1])

        # Get Node type - WDM, OCS or Not Known
        if row[-1] is None:
            node_type = 'Not Known'
        else:
            node_type = row[-1][row[-1].find('-') + 1:]

        # Get the location information
        location = row[0][:row[0].find('/')]

        # Get the SID (System ID) information
        sid = row[0][row[0].find('/') + 1 :]

        # Get the PSS Type information
        if row[-1] is None:
            pss_type = 'Not Known'
        else:
            pss_type = row[-1]

        nw_summ_sh_data_dict.update({hcid:(hcid, node_type, location, sid, pss_type)})

    # Now check every row in node coverage for missing node_type and missing pss_type
    for r_idx, row in enumerate(node_coverage_sheet.iter_rows(min_row=2, values_only=True), start=2):
        hc_id = row[0]
        
        # Only update if HC ID exists in network summary data
        if hc_id in nw_summ_sh_data_dict:
            if row[1] not in {'WDM', 'OCS'}:
                node_coverage_sheet.cell(row=r_idx, column=2).value = nw_summ_sh_data_dict[hc_id][1]

            if row[4] in {'Not Known', 'Not Available'}:
                node_coverage_sheet.cell(row=r_idx, column=5).value = nw_summ_sh_data_dict[hc_id][4]
        else:
            print(f"Warning: HC ID {hc_id} found in NODE COVERAGE but not in Network Summary sheet")

    return nw_summ_sh_data_dict

def add_new_nodes(nw_summ_sh_data_dict, node_coverage_sheet):
    nw_summ_set = set(nw_summ_sh_data_dict.values())

    node_coverage_set = set()
    # Get the existing node_coverage_sheet data
    for row in node_coverage_sheet.iter_rows(min_row=2, values_only=True):
        node_coverage_set.add(row[:5])

    new_nodes = nw_summ_set - node_coverage_set

    padding =[]
    padding_length = node_coverage_sheet.max_column - 5
    for i in range(padding_length):
        padding.append('Not Added')

    if new_nodes:
        for node in new_nodes:
            node_coverage_sheet.append(tuple(list(node) + padding))

    return



