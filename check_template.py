#!/usr/bin/env python3
"""
Script to check Template NODE COVERAGE sheet to understand expected structure
"""

import openpyxl as opxl
from pathlib import Path

def check_template_node_coverage():
    script_dir = Path("Script")
    template_file = script_dir / "Template_HC_Issues_Tracker.xlsx"
    
    if not template_file.exists():
        print(f"❌ Template file not found: {template_file}")
        return
    
    print(f"Checking Template NODE COVERAGE sheet...")
    
    try:
        wb = opxl.load_workbook(template_file)
        
        if "NODE COVERAGE" in wb.sheetnames:
            nc_sheet = wb["NODE COVERAGE"]
            print(f"✅ NODE COVERAGE sheet found in template")
            print(f"Max row: {nc_sheet.max_row}, Max col: {nc_sheet.max_column}")
            
            # Check the header row
            header_row = [cell.value for cell in nc_sheet[1]]
            print(f"Header row: {header_row}")
            
            # Check if there are template data rows
            if nc_sheet.max_row > 1:
                print(f"\n✅ Template has {nc_sheet.max_row - 1} data rows")
                
                # Show first few data rows
                for i in range(2, min(6, nc_sheet.max_row + 1)):
                    row_data = [cell.value for cell in nc_sheet[i]]
                    print(f"Row {i}: {row_data}")
                    
                print(f"\n... (showing first 4 rows only)")
                print(f"Total template nodes: {nc_sheet.max_row - 1}")
                
            else:
                print("❌ Template NODE COVERAGE sheet is also empty!")
        
        else:
            print("❌ NODE COVERAGE sheet not found in template!")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Error reading template file: {e}")

if __name__ == "__main__":
    check_template_node_coverage()
