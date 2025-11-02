#!/usr/bin/env python3
"""
Diagnostic script to troubleshoot chart generation and node coverage issues
This script checks:
1. Whether chart PNG files are being generated
2. Node coverage data population in Excel files
3. Chart embedding in summary sheets
"""

import openpyxl as opxl
from pathlib import Path
import os
import sys

def check_output_files(script_dir):
    """Check what files are generated in the output directory"""
    output_dir = script_dir / "output"
    
    print("=== CHECKING OUTPUT DIRECTORY ===")
    print(f"Output directory: {output_dir}")
    
    if not output_dir.exists():
        print("❌ Output directory does not exist!")
        return
    
    print(f"✅ Output directory exists")
    
    # List all files in output directory
    files = list(output_dir.glob("*"))
    print(f"\nFiles in output directory ({len(files)} total):")
    
    chart_files = []
    excel_files = []
    
    for file in files:
        if file.suffix.lower() == '.png':
            chart_files.append(file)
            print(f"  📊 {file.name}")
        elif file.suffix.lower() == '.xlsx':
            excel_files.append(file)
            print(f"  📋 {file.name}")
        else:
            print(f"  📄 {file.name}")
    
    print(f"\n📊 Chart files found: {len(chart_files)}")
    print(f"📋 Excel files found: {len(excel_files)}")
    
    return chart_files, excel_files

def check_chart_files(chart_files):
    """Check if the expected chart files exist"""
    print("\n=== CHECKING CHART FILES ===")
    
    hc_issues_charts = [f for f in chart_files if 'Node Coverage' not in f.name]
    node_coverage_charts = [f for f in chart_files if 'Node Coverage' in f.name]
    
    print(f"HC Issues charts: {len(hc_issues_charts)}")
    for chart in hc_issues_charts:
        print(f"  - {chart.name} ({chart.stat().st_size} bytes)")
    
    print(f"Node Coverage charts: {len(node_coverage_charts)}")
    for chart in node_coverage_charts:
        print(f"  - {chart.name} ({chart.stat().st_size} bytes)")
    
    if len(hc_issues_charts) == 0:
        print("❌ No HC Issues charts found!")
    
    if len(node_coverage_charts) == 0:
        print("❌ No Node Coverage charts found!")
    
    return hc_issues_charts, node_coverage_charts

def check_excel_summary_sheet(excel_files):
    """Check the summary sheet in Excel files for chart data and embedded images"""
    print("\n=== CHECKING EXCEL SUMMARY SHEETS ===")
    
    for excel_file in excel_files:
        if "Issues_Tracker" in excel_file.name:
            print(f"\nChecking: {excel_file.name}")
            
            try:
                wb = opxl.load_workbook(excel_file)
                
                if "Summary" in wb.sheetnames:
                    summary_sheet = wb["Summary"]
                    print(f"  ✅ SUMMARY sheet exists")
                    
                    # Check for chart data in summary sheet
                    print("  📊 Checking chart data cells:")
                    
                    # HC Issues data (G20-G22 based on main.py)
                    print(f"    G20 (Total OPEN cases): {summary_sheet['G20'].value}")
                    print(f"    G21 (New cases): {summary_sheet['G21'].value}")
                    print(f"    G22 (Closed cases): {summary_sheet['G22'].value}")
                    
                    # Node Coverage data (B20-B23)
                    print(f"    B20 (Total nodes): {summary_sheet['B20'].value}")
                    print(f"    B21 (Nodes covered): {summary_sheet['B21'].value}")
                    print(f"    B22 (Nodes not covered): {summary_sheet['B22'].value}")
                    print(f"    B23 (Nodes HC not run properly): {summary_sheet['B23'].value}")
                    
                    # Check for embedded images
                    print(f"  🖼️  Embedded images: {len(summary_sheet._images)}")
                    for i, img in enumerate(summary_sheet._images):
                        anchor_cell = f"{img.anchor._from.col_idx}{img.anchor._from.row_idx}" if hasattr(img.anchor, '_from') else "Unknown"
                        print(f"    Image {i+1}: anchored at {anchor_cell}")
                    
                    # Check NODE COVERAGE sheet
                    if "NODE COVERAGE" in wb.sheetnames:
                        nc_sheet = wb["NODE COVERAGE"]
                        print(f"  ✅ NODE COVERAGE sheet exists")
                        print(f"    Max row: {nc_sheet.max_row}, Max col: {nc_sheet.max_column}")
                        
                        # Check if there's data beyond the header
                        if nc_sheet.max_row > 1:
                            print(f"    Sample data (row 2): {[cell.value for cell in nc_sheet[2]]}")
                        else:
                            print("    ❌ No data rows in NODE COVERAGE sheet!")
                    else:
                        print("  ❌ NODE COVERAGE sheet missing!")
                        
                    wb.close()
                else:
                    print(f"  ❌ No SUMMARY sheet found")
                    
            except Exception as e:
                print(f"  ❌ Error reading Excel file: {e}")

def main():
    print("🔍 DIAGNOSTIC SCRIPT FOR CHART GENERATION AND NODE COVERAGE ISSUES")
    print("=" * 70)
    
    # Get current script directory
    script_dir = Path(__file__).parent / "Script"
    
    if not script_dir.exists():
        print("❌ Script directory not found!")
        return
    
    print(f"Script directory: {script_dir}")
    
    # Check output files
    chart_files, excel_files = check_output_files(script_dir)
    
    # Check chart files
    if chart_files:
        hc_charts, nc_charts = check_chart_files(chart_files)
    
    # Check Excel files
    if excel_files:
        check_excel_summary_sheet(excel_files)
    
    print("\n=== DIAGNOSTIC COMPLETE ===")
    print("\nPOSSIBLE ISSUES TO CHECK:")
    print("1. If no chart PNG files are generated, the chart functions might be failing")
    print("2. If chart files exist but aren't embedded, the embed functions might be failing")
    print("3. If Node Coverage data is blank/zero, check the input inventory CSV file")
    print("4. If only one chart appears, check if both chart files are generated and embedded")

if __name__ == "__main__":
    main()
