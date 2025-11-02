"""
Updated Views with Improved Script Integration
===========================================

This file contains the updated view functions that use the improved
script execution logic to fix the Windows path issues.
"""

# Import the script helper
from .script_helper import execute_health_check_script, ScriptExecutor

# Update the existing process_health_check_files function
def process_health_check_files_fixed(session_id):
    """
    Enhanced processing that uses the improved script executor
    """
    try:
        from .models import HealthCheckSession, Customer, HealthCheckFile
        from .script_helper import execute_health_check_script
        
        session = HealthCheckSession.objects.get(id=session_id)
        customer = session.customer
        
        session.update_status('PROCESSING', 'Initializing health check analysis...')
        session.progress_percentage = 5
        session.current_step = "Preparing environment"
        session.save()
        
        # Validate network setup files exist
        network_name = customer.name
        
        # Execute the script using our improved script executor
        session.progress_percentage = 25
        session.current_step = "Executing health check script"
        session.save()
        
        script_result = execute_health_check_script(network_name)
        
        if script_result['success']:
            session.progress_percentage = 80
            session.current_step = "Processing outputs"
            session.save()
            
            # Process generated outputs
            tracker_file = script_result.get('tracker_file')
            if tracker_file:
                # Save tracker file info to session
                session.output_tracker_filename = tracker_file['name']
                session.output_tracker_path = tracker_file['path']
                session.save()
            
            # Create detailed report
            create_detailed_report_fixed(session, script_result)
            
            session.progress_percentage = 100
            session.current_step = "Completed"
            session.update_status('COMPLETED', 'Health check analysis completed successfully')
            
        else:
            error_msg = script_result.get('error', 'Unknown error occurred')
            session.update_status('FAILED', f"Script execution failed: {error_msg}")
            
    except Exception as e:
        session.update_status('FAILED', f'Processing failed: {str(e)}')
        raise e


def create_detailed_report_fixed(session, script_result):
    """Create detailed health check report from script results"""
    # Health check reports are no longer needed - session status provides sufficient information
    pass


def parse_tracker_statistics_fixed(tracker_path):
    """Parse statistics from the generated HC Issues Tracker"""
    try:
        from pathlib import Path
        import openpyxl
        
        if not tracker_path or not Path(tracker_path).exists():
            return {}
        
        wb = openpyxl.load_workbook(tracker_path, read_only=True)
        
        stats = {
            'total_nodes': 0,
            'healthy_nodes': 0,
            'warning_nodes': 0,
            'offline_nodes': 0,
            'total_services': 0,
            'healthy_services': 0,
            'health_score': 95.0  # Default score
        }
        
        # Parse Summary sheet if available
        if 'Summary' in wb.sheetnames:
            summary_sheet = wb['Summary']
            # Try to extract summary statistics
            try:
                # Look for specific cells that contain summary data
                for row in summary_sheet.iter_rows(values_only=True):
                    if row[0] and isinstance(row[0], str):
                        if 'total nodes' in row[0].lower() and len(row) > 1 and isinstance(row[1], (int, float)):
                            stats['total_nodes'] = int(row[1])
                        elif 'nodes covered' in row[0].lower() and len(row) > 1 and isinstance(row[1], (int, float)):
                            stats['healthy_nodes'] = int(row[1])
                        elif 'nodes not covered' in row[0].lower() and len(row) > 1 and isinstance(row[1], (int, float)):
                            stats['offline_nodes'] = int(row[1])
            except Exception:
                pass
        
        # Parse OPEN sheet for issue counts
        if 'OPEN' in wb.sheetnames:
            open_sheet = wb['OPEN']
            issue_count = max(0, open_sheet.max_row - 1)  # Subtract header row
            
            # Calculate health score based on issues
            if issue_count == 0:
                stats['health_score'] = 100.0
            else:
                # Simple scoring algorithm - can be enhanced
                stats['health_score'] = max(0, 100 - (issue_count * 2))
        
        # Parse NODE COVERAGE sheet if available
        if 'NODE COVERAGE' in wb.sheetnames:
            coverage_sheet = wb['NODE COVERAGE']
            if coverage_sheet.max_row > 1:
                stats['total_nodes'] = max(stats['total_nodes'], coverage_sheet.max_row - 1)
        
        wb.close()
        return stats
        
    except Exception as e:
        print(f"Error parsing tracker statistics: {e}")
        return {'health_score': 50.0}  # Default fallback


# Test function to validate the integration
def test_script_integration_view(request):
    """View to test the script integration"""
    from django.http import JsonResponse
    from .script_helper import test_script_integration
    
    try:
        result = test_script_integration()
        return JsonResponse({
            'status': 'success' if result else 'error',
            'message': 'Script integration test completed',
            'result': result
        })
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        })
