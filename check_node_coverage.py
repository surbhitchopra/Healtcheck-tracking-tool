#!/usr/bin/env python3
"""
Script to check NODE COVERAGE sheet data in detail
"""

import openpyxl as opxl
from pathlib import Path

def check_node_coverage_sheet():
    script_dir = Path("Script")
    
    # Check the main tracker file
    tracker_file = script_dir / "bsnl_west_zone_otn_HC_Issues_Tracker.xlsx"
    
    if not tracker_file.exists():
        print(f"❌ Tracker file not found: {tracker_file}")
        return
    
    print(f"Checking NODE COVERAGE sheet in: {tracker_file.name}")
    
    try:
        wb = opxl.load_workbook(tracker_file)
        
        if "NODE COVERAGE" in wb.sheetnames:
            nc_sheet = wb["NODE COVERAGE"]
            print(f"✅ NODE COVERAGE sheet found")
            print(f"Max row: {nc_sheet.max_row}, Max col: {nc_sheet.max_column}")
            
            # Check the header row
            header_row = [cell.value for cell in nc_sheet[1]]
            print(f"Header row: {header_row}")
            
            # Check if there are any data rows
            if nc_sheet.max_row > 1:
                print(f"\n✅ Data rows exist ({nc_sheet.max_row - 1} data rows)")
                
                # Show first few data rows
                for i in range(2, min(6, nc_sheet.max_row + 1)):
                    row_data = [cell.value for cell in nc_sheet[i]]
                    print(f"Row {i}: {row_data}")
                
                # Check the last column (most recent HC data)
                last_col = nc_sheet.max_column
                print(f"\nLast column data (column {last_col}):")
                print(f"Header: {nc_sheet.cell(1, last_col).value}")
                
                # Count different types of values in the last column
                covered_count = 0
                not_covered_count = 0
                not_run_properly_count = 0
                
                for i in range(2, nc_sheet.max_row + 1):
                    cell_value = nc_sheet.cell(i, last_col).value
                    cell_comment = nc_sheet.cell(i, last_col).comment
                    
                    if isinstance(cell_value, str) and cell_value == "Missing":
                        not_covered_count += 1
                    elif cell_value is not None:
                        covered_count += 1
                        if cell_comment is not None:
                            not_run_properly_count += 1
                
                print(f"  Covered nodes: {covered_count}")
                print(f"  Not covered nodes: {not_covered_count}")
                print(f"  Not run properly: {not_run_properly_count}")
                
            else:
                print("❌ NO DATA ROWS FOUND! Only header row exists.")
                print("This explains why all node coverage statistics are 0.")
        
        else:
            print("❌ NODE COVERAGE sheet not found!")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Error reading file: {e}")

if __name__ == "__main__":
    check_node_coverage_sheet()
